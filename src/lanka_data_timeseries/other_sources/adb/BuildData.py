import os
import re
import tempfile

import requests
from openpyxl import load_workbook
from utils import JSONFile, Log

from lanka_data_timeseries.constants import (
    DEFAULT_FREQUENCY_NAME,
    DIR_TMP_DATA,
)
from lanka_data_timeseries.other_sources.adb.parsers import (
    I_ROW_T_HEADER,
    SOURCE_ID,
    parse_row,
)

log = Log(__file__)


def init_dir():
    dir_output = os.path.join(
        DIR_TMP_DATA,
        "sources",
        SOURCE_ID,
    )
    if not os.path.exists(dir_output):
        os.makedirs(dir_output)
        log.debug(f"Created {dir_output}")
    return dir_output


def _get_download_url(session: requests.Session) -> str:
    DATASET_URL = "https://kidb.adb.org/economies/sri-lanka"
    response = session.get(DATASET_URL, timeout=30)
    response.raise_for_status()
    # Find the first .xlsx resource link on the dataset page
    match = re.search(
        r'href="(https://data\.adb\.org/media/\d+/download)"', response.text
    )
    if match:
        return match.group(1)
    raise ValueError(f"Could not find .xlsx download URL on {DATASET_URL}")


def download_source() -> str:
    DATASET_PAGE = "https://kidb.adb.org/economies/sri-lanka"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
    }

    session = requests.Session()
    session.headers.update(headers)

    url_download = _get_download_url(session)
    log.info(f"Resolved download URL: {url_download}")

    session.headers["Referer"] = DATASET_PAGE
    session.headers["Accept"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*"
    )
    response = session.get(url_download, timeout=60)
    response.raise_for_status()

    excel_path = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
    with open(excel_path, "wb") as f:
        f.write(response.content)

    log.info(f"Downloaded {url_download} to {excel_path}")
    return excel_path


def parse_excel(excel_path: str):
    try:
        workbook = load_workbook(excel_path)
    except Exception as e:
        log.error(e)
        return []

    worksheet = workbook.active

    i_col = 3
    year_list = []
    while True:
        year_str = worksheet.cell(row=I_ROW_T_HEADER, column=i_col).value
        i_col += 1

        if not year_str:
            break
        year = int(year_str)
        year_list.append(year)

    i_row = I_ROW_T_HEADER + 1
    indent_to_text = ["" for _ in range(5)]
    category1 = ""
    last_unit = ""
    d_list = []
    MAX_ROWS = 1_000
    while i_row < MAX_ROWS:
        d, i_row, indent_to_text, category1, last_unit = parse_row(
            worksheet, i_row, year_list, indent_to_text, category1, last_unit
        )
        if d:
            d_list.append(d)
    return d_list


def build_details(d_list, dir_output):
    file_path_set = set()
    for d in d_list:
        sub_category = d["sub_category"]
        file_path = os.path.join(
            dir_output,
            f"{SOURCE_ID}.{sub_category}.{DEFAULT_FREQUENCY_NAME}.json",
        )
        if file_path in file_path_set:
            log.error(f"Duplicate file path: {file_path}")
        file_path_set.add(file_path)
        JSONFile(file_path).write(d)


def build_data():
    dir_output = init_dir()
    excel_path = download_source()
    d_list = parse_excel(excel_path)
    build_details(d_list, dir_output)


if __name__ == "__main__":
    download_source()
